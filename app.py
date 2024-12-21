from flask import Flask, render_template, request, jsonify
from talking import talk  # Make sure 'talk' is defined correctly
from take_command import takeCommand  # Make sure 'takeCommand' is defined correctly
from openpyxl import load_workbook, Workbook
import os
from speak import speak
import pywhatkit
import serial
import time
import numpy as np

app = Flask(__name__)

# File path for storing the Excel sheet
EXCEL_FILE = 'Appointmentsbook.xlsx'

@app.route('/')
def intro():
    print("Index route hit!")  # Debug print
    return render_template('intro.html')

@app.route('/start')
def index():
    return render_template('index.html')

@app.route('/meditation')
def meditation():
    return render_template('meditation.html')  # Render the meditation page

@app.route('/aiassistant')
def aiassistant():
    return render_template('aiassistant.html')  # Render the meditation page

@app.route('/calmmeditation')
def calmmeditation():
    return render_template('calmmeditation.html')  # Render the meditation page

@app.route('/breathing')
def breathing():
    return render_template('breathing.html')  # Render the meditation page

@app.route('/manifest')
def manifest():
    return render_template('manifest.html')  # Render the meditation page

@app.route('/soulheal')
def soulheal():
    return render_template('soulheal.html')  # Render the tasks page

@app.route('/tasks')
def tasks():
    return render_template('tasks.html')  # Render the tasks page

@app.route('/courses')
def courses():
    return render_template('courses.html')  # Render the tasks page

@app.route('/mydiary')
def mydiary():
    return render_template('MyDiary.html')  # Render the meditation page

@app.route('/survey')
def survey():
    return render_template('survey.html')  # Render the meditation page

@app.route('/memories')
def memories():
    return render_template('memories.html')  # Render the tasks page

@app.route('/stats')
def stats():
    return render_template('stats.html')  # Your statistics page

@app.route('/run-pls', methods=['POST'])
def pls():
    query, emotion = takeCommand()  # Now receiving both query and emotion
    query = query.lower() if query != "None" else "none"
    response_output = ""
    
    if query != "none":
        response_output = talk(query, emotion)  # Pass emotion to talk function
        speak(response_output)
    
    return jsonify({
        "status": "success", 
        "message": "Command executed", 
        "output": response_output,
        "detected_emotion": emotion
    })



@app.route('/send-query', methods=['POST'])
def send_query():
    data = request.json
    user_query = data.get('query', '')
    output = talk(user_query)
    response = {
        'output': output
    }
    return jsonify(response)


@app.route('/send-whatsapp', methods=['POST'])
def send_whatsapp():
    try:
        # Specify the recipient's number and the message
        phone_number = "+919886093707"
        message = "Hello"
        
        # Send the message instantly
        pywhatkit.sendwhatmsg_instantly(phone_number, message)

        # Return a success response
        return jsonify({"status": "success", "message": "WhatsApp message sent!"}), 200
    except Exception as e:
        # Handle any errors
        return jsonify({"status": "error", "message": str(e)}), 500


def write_to_excel(booking_data):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Find the next available row in the spreadsheet
        next_row = ws.max_row + 1
        for col_num, entry in enumerate(booking_data, start=1):
            ws.cell(row=next_row, column=col_num, value=entry)
        
        wb.save(EXCEL_FILE)
    except FileNotFoundError:
        # If the file doesn't exist, create it and write headers
        wb = Workbook()
        ws = wb.active
        headers = ["Name", "Email", "Phone", "Purpose"]
        ws.append(headers)
        ws.append(booking_data)
        wb.save(EXCEL_FILE)

@app.route('/submit-booking', methods=['POST'])
def submit_booking():
    # Retrieve data from the form
    name = request.form.get('name', '')
    email = request.form.get('email', '')
    phone = request.form.get('phone', '')
    purpose = request.form.get('purpose', '')

    # Prepare the data for Excel
    booking_data = [name, email, phone, purpose]

    # Write data to Excel
    write_to_excel(booking_data)

    # Render the appointment confirmation page
    return render_template('appointment.html', name=name, purpose=purpose)


# Global Arduino object
arduino = None

# Function to initialize Arduino
def initialize_arduino():
    global arduino
    try:
        if arduino is None or not arduino.is_open:
            arduino = serial.Serial('COM7', 9600)
            print("Arduino Working")
    except Exception as e:
        print(f"Arduino connection error: {e}")

# RMSSD calculation
def calculate_rmssd(ibi_list):
    if len(ibi_list) < 2:
        return None
    differences = np.diff(ibi_list)
    squared_differences = differences ** 2
    rmssd = np.sqrt(np.mean(squared_differences))
    return rmssd

# Stress level interpretation based on RMSSD and BPM
def interpret_stress(rmssd, bpm):
    """
    Interpret stress level using both RMSSD and BPM.
    """
    if rmssd is None or bpm is None:
        return "Insufficient data"
    
    if bpm > 100:
        # High BPM cases
        if rmssd > 50:
            return "Low stress"  # Relaxation despite high heart rate
        elif 30 < rmssd <= 50:
            return "Moderate stress"  # Elevated heart rate with moderate relaxation
        else:
            return "High stress"  # High heart rate and low RMSSD indicate stress
    else:
        # Normal BPM cases
        if rmssd > 50:
            return "Low stress"
        elif 30 < rmssd <= 50:
            return "Moderate stress"
        else:
            return "High stress"

# Fetch sensor data from Arduino
ibi_values = []

def get_sensor_data():
    initialize_arduino()  # Ensure Arduino is initialized before reading data
    if arduino and arduino.in_waiting > 0:
        try:
            data = arduino.readline().decode('utf-8').strip()
            print(f"Received data: {data}")  # Debugging print statement
            
            # Ensure the data format is correct before processing
            if "BPM:" in data and "IBI:" in data:
                # Parse BPM and IBI values
                bpm_value = int(data.split(", ")[0].split(": ")[1])
                ibi_value = int(data.split(", ")[1].split(": ")[1])

                # Only process valid BPM and IBI values
                if bpm_value > 0 and ibi_value > 0:
                    ibi_values.append(ibi_value)
                    if len(ibi_values) > 20:
                        ibi_values.pop(0)

                    # Calculate RMSSD and interpret stress
                    rmssd = calculate_rmssd(ibi_values)
                    stress_level = interpret_stress(rmssd, bpm_value)

                    return bpm_value, rmssd, stress_level
                else:
                    print("Invalid BPM or IBI values")
            else:
                print("Data format incorrect")
        except Exception as e:
            print(f"Error processing sensor data: {e}")
    else:
        print("No data available or Arduino not connected")
    return None, None, None

# Flask routes
@app.route('/watch')
def watch():
    return render_template('watch.html')

@app.route('/watch/data')
def data():
    bpm, rmssd, stress_level = get_sensor_data()
    return jsonify({
        "bpm": bpm, #if bpm is not None else "No data",
        "rmssd": rmssd, #if rmssd is not None else "No data",
        "stress_level": stress_level #if stress_level else "No data"
    })

if __name__ == '__main__':
    print("Starting Flask app...")  # Debug print
    app.run(debug=True)
